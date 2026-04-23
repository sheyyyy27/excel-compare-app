import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Excel Subsidiary Comparison Tool", layout="wide")
st.title("Excel Subsidiary Comparison Tool")
st.caption("Display page stays as the master file. Parse page is checked for updates and new rows.")

display_file = st.file_uploader("Upload Display Sheet", type=["xlsx"])
parse_file = st.file_uploader("Upload Parse Sheet", type=["xlsx"])

DISPLAY_COLUMNS = [
    "Checked",
    "Subsidiary Name",
    "Source",
    "Parent Entity Name",
    "Incorporation Location",
    "Ownership Percentage",
    "Entity Type Code",
    "Entity SubType Code",
    "Subsidiary Comments",
    "Domain",
    "Domain Comments",
    "Address",
    "Country",
    "Changes",
]

KEY_COL = "Subsidiary Name"

UPDATE_COLUMNS = [
    "Source",
    "Parent Entity Name",
    "Incorporation Location",
    "Ownership Percentage",
    "Entity Type Code",
    "Entity SubType Code",
    "Address",
    "Country",
]

KEEP_FROM_DISPLAY = [
    "Subsidiary Comments",
    "Domain Comments",
]

DOMAIN_COL = "Domain"
CHECKED_COL = "Checked"
CHANGES_COL = "Changes"

HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")

LEGAL_EQUIVALENTS = {
    "co": "company",
    "co.": "company",
    "company": "company",
    "companies": "company",
    "corp": "corporation",
    "corp.": "corporation",
    "corporation": "corporation",
    "inc": "incorporated",
    "inc.": "incorporated",
    "incorporated": "incorporated",
    "ltd": "limited",
    "ltd.": "limited",
    "limited": "limited",
    "llc": "llc",
    "l.l.c.": "llc",
    "llp": "llp",
    "l.l.p.": "llp",
    "plc": "plc",
    "p.l.c.": "plc",
    "sa": "sa",
    "s.a.": "sa",
    "nv": "nv",
    "bv": "bv",
    "ag": "ag",
    "gmbh": "gmbh",
    "oy": "oy",
    "ab": "ab",
    "kk": "kk",
}

SOURCE_SPLIT_PATTERN = re.compile(r"\s*,\s*")


def ensure_columns(df: pd.DataFrame, expected_cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""
    return df


def clean_string(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_for_compare(value) -> str:
    text = clean_string(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_source(value) -> str:
    text = clean_string(value)
    if not text:
        return ""
    parts = [p.strip() for p in SOURCE_SPLIT_PATTERN.split(text) if p.strip()]
    parts = [re.sub(r"\s+", " ", p) for p in parts]
    return ",".join(sorted({p.lower() for p in parts}))


def normalize_name(name: str) -> str:
    text = clean_string(name).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^\w\s]", " ", text)
    words = [w for w in text.split() if w]

    normalized_words = []
    for word in words:
        normalized_words.append(LEGAL_EQUIVALENTS.get(word, word))

    return " ".join(normalized_words).strip()


def domain_already_exists(display_domain: str, parse_domain: str) -> bool:
    return normalize_for_compare(display_domain) == normalize_for_compare(parse_domain)


def values_equal(col: str, old_val, new_val) -> bool:
    if col == "Source":
        return normalize_source(old_val) == normalize_source(new_val)
    return normalize_for_compare(old_val) == normalize_for_compare(new_val)


def build_display_lookup(display_df: pd.DataFrame) -> dict[str, int]:
    lookup = {}
    for idx, row in display_df.iterrows():
        key = clean_string(row.get(KEY_COL, ""))
        if key and key not in lookup:
            lookup[key] = idx
    return lookup


def build_normalized_name_set(display_df: pd.DataFrame) -> set[str]:
    normalized_set = set()
    for _, row in display_df.iterrows():
        name = clean_string(row.get(KEY_COL, ""))
        if name:
            normalized_set.add(normalize_name(name))
    return normalized_set


def add_change_message(existing: str, new_msg: str) -> str:
    existing = clean_string(existing)
    new_msg = clean_string(new_msg)

    if not existing:
        return new_msg

    existing_parts = [p.strip() for p in existing.split(" | ") if p.strip()]
    if new_msg in existing_parts:
        return existing

    return f"{existing} | {new_msg}"


def compare_and_update(display_df: pd.DataFrame, parse_df: pd.DataFrame):
    display_df = ensure_columns(display_df, DISPLAY_COLUMNS[:-1])
    parse_df = ensure_columns(parse_df, DISPLAY_COLUMNS[:-1])

    # Add Changes column if missing in Display
    if CHANGES_COL not in display_df.columns:
        display_df[CHANGES_COL] = ""

    output_df = display_df.copy()

    changed_cells = []  # (excel_row, excel_col)
    display_lookup = build_display_lookup(output_df)
    normalized_display_names = build_normalized_name_set(output_df)

    summary = {
        "updated_rows": 0,
        "new_rows": 0,
        "possible_duplicates": 0,
        "domains_added": 0,
        "different_domains_logged": 0,
        "duplicate_domains_logged": 0,
    }

    for _, parse_row in parse_df.iterrows():
        parse_name = clean_string(parse_row.get(KEY_COL, ""))
        if not parse_name:
            continue

        if parse_name in display_lookup:
            row_idx = display_lookup[parse_name]
            updated_cols = []

            # Compare update columns
            for col in UPDATE_COLUMNS:
                old_val = output_df.at[row_idx, col] if col in output_df.columns else ""
                new_val = parse_row.get(col, "")

                if not values_equal(col, old_val, new_val):
                    output_df.at[row_idx, col] = new_val
                    updated_cols.append(col)

                    excel_row = row_idx + 2
                    excel_col = output_df.columns.get_loc(col) + 1
                    changed_cells.append((excel_row, excel_col))

            # Domain logic
            display_domain = clean_string(output_df.at[row_idx, DOMAIN_COL])
            parse_domain = clean_string(parse_row.get(DOMAIN_COL, ""))

            if parse_domain:
                if not display_domain:
                    output_df.at[row_idx, DOMAIN_COL] = parse_domain
                    excel_row = row_idx + 2
                    excel_col = output_df.columns.get_loc(DOMAIN_COL) + 1
                    changed_cells.append((excel_row, excel_col))
                    output_df.at[row_idx, CHANGES_COL] = add_change_message(
                        output_df.at[row_idx, CHANGES_COL],
                        f"Domain added: {parse_domain}",
                    )
                    summary["domains_added"] += 1
                elif domain_already_exists(display_domain, parse_domain):
                    output_df.at[row_idx, CHANGES_COL] = add_change_message(
                        output_df.at[row_idx, CHANGES_COL],
                        f"{parse_domain} already in Domain column",
                    )
                    summary["duplicate_domains_logged"] += 1
                else:
                    output_df.at[row_idx, CHANGES_COL] = add_change_message(
                        output_df.at[row_idx, CHANGES_COL],
                        f"Different Parse domain found: {parse_domain}",
                    )
                    summary["different_domains_logged"] += 1

            if updated_cols:
                output_df.at[row_idx, CHANGES_COL] = add_change_message(
                    output_df.at[row_idx, CHANGES_COL],
                    f"Updated: {', '.join(updated_cols)}",
                )
                summary["updated_rows"] += 1

        else:
            # New row from Parse
            new_row = {col: "" for col in output_df.columns}

            new_row[CHECKED_COL] = False
            new_row[KEY_COL] = parse_name

            for col in output_df.columns:
                if col in [CHECKED_COL, CHANGES_COL, "Subsidiary Comments", "Domain Comments"]:
                    continue
                if col in parse_df.columns:
                    new_row[col] = parse_row.get(col, "")

            # Preserve blank comments for new rows
            new_row["Subsidiary Comments"] = ""
            new_row["Domain Comments"] = ""

            normalized_parse_name = normalize_name(parse_name)
            if normalized_parse_name in normalized_display_names:
                new_row[CHANGES_COL] = "New subsidiary added from Parse | Possible duplicate (legal-name variation)"
                summary["possible_duplicates"] += 1
            else:
                new_row[CHANGES_COL] = "New subsidiary added from Parse"

            output_df = pd.concat([output_df, pd.DataFrame([new_row])], ignore_index=True)
            new_row_idx = len(output_df) - 1
            display_lookup[parse_name] = new_row_idx
            normalized_display_names.add(normalized_parse_name)
            summary["new_rows"] += 1

            # Highlight imported cells for new row
            highlight_cols = [
                "Checked",
                "Subsidiary Name",
                "Source",
                "Parent Entity Name",
                "Incorporation Location",
                "Ownership Percentage",
                "Entity Type Code",
                "Entity SubType Code",
                "Domain",
                "Address",
                "Country",
                "Changes",
            ]
            for col in highlight_cols:
                if col in output_df.columns:
                    excel_row = new_row_idx + 2
                    excel_col = output_df.columns.get_loc(col) + 1
                    changed_cells.append((excel_row, excel_col))

    # Make sure final column order is correct
    output_df = ensure_columns(output_df, DISPLAY_COLUMNS)
    output_df = output_df[DISPLAY_COLUMNS]

    return output_df, changed_cells, summary


def create_excel_file(df: pd.DataFrame, changed_cells: list[tuple[int, int]]) -> BytesIO:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Output")
        ws = writer.book["Output"]

        for row_num, col_num in changed_cells:
            ws.cell(row=row_num, column=col_num).fill = HIGHLIGHT_FILL

        # Freeze header
        ws.freeze_panes = "A2"

        # Basic width adjustment
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)

    output.seek(0)
    return output


if display_file and parse_file:
    try:
        display_df = pd.read_excel(display_file)
        parse_df = pd.read_excel(parse_file)

        required_display_cols = [
            "Checked",
            "Subsidiary Name",
            "Source",
            "Parent Entity Name",
            "Incorporation Location",
            "Ownership Percentage",
            "Entity Type Code",
            "Entity SubType Code",
            "Subsidiary Comments",
            "Domain",
            "Domain Comments",
            "Address",
            "Country",
        ]

        required_parse_cols = [
            "Subsidiary Name",
            "Source",
            "Parent Entity Name",
            "Incorporation Location",
            "Ownership Percentage",
            "Entity Type Code",
            "Entity SubType Code",
            "Domain",
            "Address",
            "Country",
        ]

        missing_display = [c for c in required_display_cols if c not in display_df.columns]
        missing_parse = [c for c in required_parse_cols if c not in parse_df.columns]

        if missing_display:
            st.error(f"Display file is missing columns: {', '.join(missing_display)}")
        elif missing_parse:
            st.error(f"Parse file is missing columns: {', '.join(missing_parse)}")
        else:
            result_df, changed_cells, summary = compare_and_update(display_df, parse_df)

            col1, col2, col3 = st.columns(3)
            col1.metric("Updated existing rows", summary["updated_rows"])
            col2.metric("New rows added", summary["new_rows"])
            col3.metric("Possible duplicates logged", summary["possible_duplicates"])

            col4, col5, col6 = st.columns(3)
            col4.metric("Domains added", summary["domains_added"])
            col5.metric("Different domains logged", summary["different_domains_logged"])
            col6.metric("Duplicate domains logged", summary["duplicate_domains_logged"])

            st.subheader("Output Preview")
            st.dataframe(result_df, use_container_width=True)

            excel_output = create_excel_file(result_df, changed_cells)

            st.download_button(
                label="Download Compared Excel File",
                data=excel_output,
                file_name="comparison_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"Error: {e}")
